import logging
import os
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    filters, ContextTypes, ConversationHandler
)
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# Excel file paths
FLIGHTS_FILE = 'flights.xlsx'
BOOKINGS_FILE = 'bookings.xlsx'

# Admin user IDs (replace with actual admin Telegram IDs)
ADMIN_USERS = [123456789]

# Conversation states
SELECTING_DATE, SELECTING_FLIGHT, CONFIRMING_BOOKING, ADMIN_ADDING_FLIGHT = range(4)

# Callback data prefixes
DATE_PREFIX = "date:"
FLIGHT_PREFIX = "flight:"
CONFIRM_PREFIX = "confirm:"
CANCEL_PREFIX = "cancel:"
ADMIN_PREFIX = "admin:"

# Create sample flights Excel file with realistic data
def create_flights_excel():
    """Create a sample flights Excel file with realistic flight data."""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Flights"
    
    # Add headers with styling
    headers = ["Date", "Time", "Flight Number", "Departure", "Destination", "Capacity", "Booked"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Generate sample flight data
    # Use the next 7 days for flights
    start_date = datetime.now().date()
    flight_data = []
    
    # Flight routes
    routes = [
        {"flight": "FL101", "departure": "New York", "destination": "London", "capacity": 120},
        {"flight": "FL102", "departure": "London", "destination": "New York", "capacity": 120},
        {"flight": "FL203", "departure": "Paris", "destination": "Tokyo", "capacity": 180},
        {"flight": "FL204", "departure": "Tokyo", "destination": "Paris", "capacity": 180},
        {"flight": "FL305", "departure": "Dubai", "destination": "Sydney", "capacity": 150},
        {"flight": "FL306", "departure": "Sydney", "destination": "Dubai", "capacity": 150},
        {"flight": "FL407", "departure": "Singapore", "destination": "San Francisco", "capacity": 200},
        {"flight": "FL408", "departure": "San Francisco", "destination": "Singapore", "capacity": 200}
    ]
    
    # Flight times
    times = ["06:30", "08:45", "11:15", "13:30", "16:00", "19:45", "22:30"]
    
    # Generate flights for each day
    row_num = 2  # Start from row 2 (after header)
    for day_offset in range(7):
        flight_date = start_date + timedelta(days=day_offset)
        date_str = flight_date.strftime("%Y-%m-%d")
        
        # Add different flights on different days for variety
        daily_routes = routes[day_offset % 4:day_offset % 4 + 4]
        
        for route in daily_routes:
            # Add 2-3 time slots per route per day
            for time in times[day_offset % 3:day_offset % 3 + 3]:
                # Add data to worksheet
                ws.cell(row=row_num, column=1).value = date_str
                ws.cell(row=row_num, column=2).value = time
                ws.cell(row=row_num, column=3).value = route["flight"]
                ws.cell(row=row_num, column=4).value = route["departure"]
                ws.cell(row=row_num, column=5).value = route["destination"]
                ws.cell(row=row_num, column=6).value = route["capacity"]
                ws.cell(row=row_num, column=7).value = 0  # Start with 0 bookings
                
                # Center align all cells except departure and destination
                for col in [1, 2, 3, 6, 7]:
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
                
                row_num += 1
    
    # Adjust column widths
    column_widths = {
        "A": 12,  # Date
        "B": 10,  # Time
        "C": 15,  # Flight Number
        "D": 20,  # Departure
        "E": 20,  # Destination
        "F": 10,  # Capacity
        "G": 10   # Booked
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add border to all cells
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Save the workbook
    wb.save(FLIGHTS_FILE)
    logger.info(f"Created flights Excel file: {FLIGHTS_FILE}")

# Create empty bookings Excel file
def create_bookings_excel():
    """Create an empty bookings Excel file with proper structure."""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    
    # Add headers with styling
    headers = ["Date", "Time", "Flight Number", "User ID", "Username", "Booking Time"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    column_widths = {
        "A": 12,  # Date
        "B": 10,  # Time
        "C": 15,  # Flight Number
        "D": 12,  # User ID
        "E": 20,  # Username
        "F": 20   # Booking Time
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add border to header row
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.border = thin_border
    
    # Save the workbook
    wb.save(BOOKINGS_FILE)
    logger.info(f"Created bookings Excel file: {BOOKINGS_FILE}")

# Modified ensure_excel_files_exist function to use our new templates
def ensure_excel_files_exist():
    """Ensure Excel files exist, creating them with sample data if not."""
    if not os.path.exists(FLIGHTS_FILE):
        create_flights_excel()
    
    if not os.path.exists(BOOKINGS_FILE):
        create_bookings_excel()

# Command to manually recreate the Excel files (admin only)
async def recreate_excel_files(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recreate Excel files with sample data (admin only)."""
    # Check if user is admin
    if update.effective_user.id not in ADMIN_USERS:
        await update.message.reply_text("This command is only available to administrators.")
        return
    
    try:
        # Recreate files
        create_flights_excel()
        create_bookings_excel()
        await update.message.reply_text("✅ Excel files recreated successfully with sample data.")
    except Exception as e:
        logger.error(f"Error recreating Excel files: {e}")
        await update.message.reply_text(f"Error recreating Excel files: {e}")

# Command handlers
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send a message when the command /start is issued."""
    ensure_excel_files_exist()
    
    message = (
        "Welcome to the Flight Booking Bot!\n\n"
        "Commands:\n"
        "/book - Book a flight\n"
        "/my_bookings - View your bookings\n"
        "/cancel_booking - Cancel a booking\n"
        "/available - Show available flights\n"
    )
    
    # Add admin commands if user is admin
    if update.effective_user.id in ADMIN_USERS:
        message += "\nAdmin Commands:\n"
        message += "/add_flight - Add a new flight\n"
        message += "/remove_flight - Remove a flight\n"
        message += "/download_bookings - Download bookings data\n"
        message += "/recreate_excel - Recreate Excel files with sample data\n"
    
    await update.message.reply_text(message)

async def book_flight_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start the booking process by showing available dates."""
    ensure_excel_files_exist()
    
    # Read the flights data
    try:
        df = pd.read_excel(FLIGHTS_FILE)
        if df.empty:
            await update.message.reply_text("No flights are currently available for booking.")
            return ConversationHandler.END
        
        # Get unique dates with available seats
        df['Date'] = pd.to_datetime(df['Date']).dt.date  # Convert to date objects
        df['Available'] = df['Capacity'] - df['Booked']
        available_flights = df[df['Available'] > 0]
        
        if available_flights.empty:
            await update.message.reply_text("Sorry, all flights are fully booked.")
            return ConversationHandler.END
        
        unique_dates = sorted(available_flights['Date'].unique())
        
        # Create keyboard with dates
        keyboard = []
        for date in unique_dates:
            date_str = date.strftime("%Y-%m-%d")
            keyboard.append([InlineKeyboardButton(date_str, callback_data=f"{DATE_PREFIX}{date_str}")])
        
        # Add cancel button
        keyboard.append([InlineKeyboardButton("Cancel", callback_data="cancel")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Please select a date:", reply_markup=reply_markup)
        
        return SELECTING_DATE
    
    except Exception as e:
        logger.error(f"Error in book_flight_start: {e}")
        await update.message.reply_text(f"Error retrieving flight dates: {e}")
        return ConversationHandler.END

async def select_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle date selection and show available flights for that date."""
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.edit_message_text("Booking cancelled.")
        return ConversationHandler.END
    
    selected_date = query.data.replace(DATE_PREFIX, "")
    context.user_data['selected_date'] = selected_date
    
    try:
        # Read flights for the selected date
        df = pd.read_excel(FLIGHTS_FILE)
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        df['Available'] = df['Capacity'] - df['Booked']
        
        # Filter by date and available seats
        date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
        available_flights = df[(df['Date'] == date_obj) & (df['Available'] > 0)]
        
        if available_flights.empty:
            await query.edit_message_text(f"No available flights on {selected_date}. Please try another date.")
            return ConversationHandler.END
        
        # Create keyboard with flights
        keyboard = []
        for _, flight in available_flights.iterrows():
            flight_info = f"{flight['Time']} - {flight['Flight Number']} - {flight['Departure']} to {flight['Destination']} ({flight['Available']} seats)"
            flight_data = f"{flight['Time']}|{flight['Flight Number']}"
            keyboard.append([InlineKeyboardButton(flight_info, callback_data=f"{FLIGHT_PREFIX}{flight_data}")])
        
        # Add back and cancel buttons
        keyboard.append([
            InlineKeyboardButton("« Back", callback_data="back_to_dates"),
            InlineKeyboardButton("Cancel", callback_data="cancel")
        ])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"Available flights on {selected_date}. Please select a flight:",
            reply_markup=reply_markup
        )
        
        return SELECTING_FLIGHT
    
    except Exception as e:
        logger.error(f"Error in select_date: {e}")
        await query.edit_message_text(f"Error retrieving flights: {e}")
        return ConversationHandler.END

async def select_flight(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle flight selection and ask for confirmation."""
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.edit_message_text("Booking cancelled.")
        return ConversationHandler.END
    
    if query.data == "back_to_dates":
        return await book_flight_start(update, context)
    
    selected_flight = query.data.replace(FLIGHT_PREFIX, "")
    flight_time, flight_number = selected_flight.split("|")
    
    context.user_data['selected_time'] = flight_time
    context.user_data['selected_flight'] = flight_number
    
    try:
        # Get flight details
        df = pd.read_excel(FLIGHTS_FILE)
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        date_obj = datetime.strptime(context.user_data['selected_date'], "%Y-%m-%d").date()
        
        flight_details = df[
            (df['Date'] == date_obj) & 
            (df['Time'] == flight_time) & 
            (df['Flight Number'] == flight_number)
        ].iloc[0]
        
        # Create confirmation message
        confirmation_msg = (
            f"Please confirm your booking:\n\n"
            f"Date: {context.user_data['selected_date']}\n"
            f"Time: {flight_time}\n"
            f"Flight: {flight_number}\n"
            f"Route: {flight_details['Departure']} to {flight_details['Destination']}"
        )
        
        # Create confirmation keyboard
        keyboard = [
            [
                InlineKeyboardButton("Confirm", callback_data=f"{CONFIRM_PREFIX}confirm"),
                InlineKeyboardButton("Cancel", callback_data="cancel")
            ]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(confirmation_msg, reply_markup=reply_markup)
        
        return CONFIRMING_BOOKING
    
    except Exception as e:
        logger.error(f"Error in select_flight: {e}")
        await query.edit_message_text(f"Error retrieving flight details: {e}")
        return ConversationHandler.END

async def confirm_booking(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle booking confirmation and save the booking."""
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.edit_message_text("Booking cancelled.")
        return ConversationHandler.END
    
    try:
        # Get booking details from context
        date_str = context.user_data['selected_date']
        time_str = context.user_data['selected_time']
        flight_number = context.user_data['selected_flight']
        user_id = update.effective_user.id
        username = update.effective_user.username or update.effective_user.first_name
        booking_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Check if user already has this booking
        bookings_df = pd.read_excel(BOOKINGS_FILE)
        existing_booking = bookings_df[
            (bookings_df['User ID'] == user_id) & 
            (bookings_df['Date'] == date_str) & 
            (bookings_df['Flight Number'] == flight_number)
        ]
        
        if not existing_booking.empty:
            await query.edit_message_text("You have already booked this flight.")
            return ConversationHandler.END
        
        # Check if flight is still available
        flights_df = pd.read_excel(FLIGHTS_FILE)
        flights_df['Date'] = pd.to_datetime(flights_df['Date']).dt.strftime("%Y-%m-%d")
        flight_row = flights_df[
            (flights_df['Date'] == date_str) & 
            (flights_df['Time'] == time_str) & 
            (flights_df['Flight Number'] == flight_number)
        ]
        
        if flight_row.empty:
            await query.edit_message_text("This flight is no longer available.")
            return ConversationHandler.END
        
        flight_index = flight_row.index[0]
        if flights_df.at[flight_index, 'Booked'] >= flights_df.at[flight_index, 'Capacity']:
            await query.edit_message_text("Sorry, this flight is now fully booked.")
            return ConversationHandler.END
        
        # Add the booking
        new_booking = pd.DataFrame({
            'Date': [date_str],
            'Time': [time_str],
            'Flight Number': [flight_number],
            'User ID': [user_id],
            'Username': [username],
            'Booking Time': [booking_time]
        })
        
        bookings_df = pd.concat([bookings_df, new_booking], ignore_index=True)
        bookings_df.to_excel(BOOKINGS_FILE, index=False)
        
        # Update flight booking count
        flights_df.at[flight_index, 'Booked'] += 1
        flights_df.to_excel(FLIGHTS_FILE, index=False)
        
        await query.edit_message_text(
            f"✅ Booking confirmed!\n\n"
            f"Date: {date_str}\n"
            f"Time: {time_str}\n"
            f"Flight: {flight_number}\n\n"
            f"You can view your bookings with /my_bookings"
        )
        
        return ConversationHandler.END
    
    except Exception as e:
        logger.error(f"Error in confirm_booking: {e}")
        await query.edit_message_text(f"Error processing booking: {e}")
        return ConversationHandler.END

async def my_bookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show user's bookings."""
    ensure_excel_files_exist()
    user_id = update.effective_user.id
    
    try:
        # Read bookings
        bookings_df = pd.read_excel(BOOKINGS_FILE)
        user_bookings = bookings_df[bookings_df['User ID'] == user_id]
        
        if user_bookings.empty:
            await update.message.reply_text("You don't have any bookings.")
            return
        
        # Get flight details
        flights_df = pd.read_excel(FLIGHTS_FILE)
        flights_df['Date'] = pd.to_datetime(flights_df['Date']).dt.strftime("%Y-%m-%d")
        
        # Create message with bookings
        message = "Your bookings:\n\n"
        
        for index, booking in user_bookings.iterrows():
            flight_info = flights_df[
                (flights_df['Date'] == booking['Date']) & 
                (flights_df['Flight Number'] == booking['Flight Number'])
            ]
            
            if not flight_info.empty:
                flight = flight_info.iloc[0]
                message += (
                    f"{index + 1}. Date: {booking['Date']}\n"
                    f"   Time: {booking['Time']}\n"
                    f"   Flight: {booking['Flight Number']}\n"
                    f"   Route: {flight['Departure']} to {flight['Destination']}\n\n"
                )
            else:
                message += (
                    f"{index + 1}. Date: {booking['Date']}\n"
                    f"   Time: {booking['Time']}\n"
                    f"   Flight: {booking['Flight Number']}\n\n"
                )
        
        await update.message.reply_text(message)
    
    except Exception as e:
        logger.error(f"Error in my_bookings: {e}")
        await update.message.reply_text(f"Error retrieving bookings: {e}")

async def cancel_booking_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start the process of cancelling a booking."""
    ensure_excel_files_exist()
    user_id = update.effective_user.id
    
    try:
        # Read bookings
        bookings_df = pd.read_excel(BOOKINGS_FILE)
        user_bookings = bookings_df[bookings_df['User ID'] == user_id]
        
        if user_bookings.empty:
            await update.message.reply_text("You don't have any bookings to cancel.")
            return ConversationHandler.END
        
        # Read flights for additional info
        flights_df = pd.read_excel(FLIGHTS_FILE)
        flights_df['Date'] = pd.to_datetime(flights_df['Date']).dt.strftime("%Y-%m-%d")
        
        # Create keyboard with bookings
        keyboard = []
        
        for index, booking in user_bookings.iterrows():
            flight_info = flights_df[
                (flights_df['Date'] == booking['Date']) & 
                (flights_df['Flight Number'] == booking['Flight Number'])
            ]
            
            if not flight_info.empty:
                flight = flight_info.iloc[0]
                button_text = f"{booking['Date']} - {booking['Time']} - {flight['Departure']} to {flight['Destination']}"
            else:
                button_text = f"{booking['Date']} - {booking['Time']} - {booking['Flight Number']}"
            
            callback_data = f"{CANCEL_PREFIX}{index}"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])
        
        # Add cancel button
        keyboard.append([InlineKeyboardButton("Cancel", callback_data="cancel")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Select a booking to cancel:", reply_markup=reply_markup)
        
        # Store bookings dataframe in context for later use
        context.user_data['bookings_df'] = bookings_df
        
        return CONFIRMING_BOOKING
    
    except Exception as e:
        logger.error(f"Error in cancel_booking_start: {e}")
        await update.message.reply_text(f"Error retrieving bookings: {e}")
        return ConversationHandler.END

async def cancel_booking_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle booking cancellation."""
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.edit_message_text("Cancellation aborted.")
        return ConversationHandler.END
    
    booking_index = int(query.data.replace(CANCEL_PREFIX, ""))
    
    try:
        # Get booking details
        bookings_df = context.user_data['bookings_df']
        user_id = update.effective_user.id
        user_bookings = bookings_df[bookings_df['User ID'] == user_id]
        booking = user_bookings.iloc[booking_index]
        
        # Remove booking
        booking_row_index = user_bookings.index[booking_index]
        bookings_df = bookings_df.drop(booking_row_index)
        bookings_df.to_excel(BOOKINGS_FILE, index=False)
        
        # Update flight booking count
        flights_df = pd.read_excel(FLIGHTS_FILE)
        flights_df['Date'] = pd.to_datetime(flights_df['Date']).dt.strftime("%Y-%m-%d")
        
        flight_row = flights_df[
            (flights_df['Date'] == booking['Date']) & 
            (flights_df['Time'] == booking['Time']) & 
            (flights_df['Flight Number'] == booking['Flight Number'])
        ]
        
        if not flight_row.empty:
            flight_index = flight_row.index[0]
            flights_df.at[flight_index, 'Booked'] = max(0, flights_df.at[flight_index, 'Booked'] - 1)
            flights_df.to_excel(FLIGHTS_FILE, index=False)
        
        await query.edit_message_text(
            f"✅ Booking cancelled successfully:\n\n"
            f"Date: {booking['Date']}\n"
            f"Time: {booking['Time']}\n"
            f"Flight: {booking['Flight Number']}"
        )
        
        return ConversationHandler.END
    
    except Exception as e:
        logger.error(f"Error in cancel_booking_confirm: {e}")
        await query.edit_message_text(f"Error cancelling booking: {e}")
        return ConversationHandler.END

async def show_available_flights(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show available flights for the coming days."""
    ensure_excel_files_exist()
    
    try:
        # Read flights data
        df = pd.read_excel(FLIGHTS_FILE)
        if df.empty:
            await update.message.reply_text("No flights are currently scheduled.")
            return
        
        # Convert date and calculate available seats
        df['Date'] = pd.to_datetime(df['Date'])
        df['Available'] = df['Capacity'] - df['Booked']
        
        # Filter for future flights with available seats
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        available_flights = df[(df['Date'] >= today) & (df['Available'] > 0)]
        
        if available_flights.empty:
            await update.message.reply_text("No available flights found for the coming days.")
            return
        
        # Sort by date and time
        available_flights = available_flights.sort_values(by=['Date', 'Time'])
        
        # Group by date
        message = "Available flights:\n\n"
        current_date = None
        
        for _, flight in available_flights.iterrows():
            flight_date = flight['Date'].strftime("%Y-%m-%d")
            
            if flight_date != current_date:
                message += f"📅 {flight_date}:\n"
                current_date = flight_date
            
            message += (
                f"  • {flight['Time']} - Flight {flight['Flight Number']}\n"
                f"    {flight['Departure']} to {flight['Destination']}\n"
                f"    Available seats: {flight['Available']}\n\n"
            )
        
        message += "Use /book to book a flight."
        await update.message.reply_text(message)
    
    except Exception as e:
        logger.error(f"Error in show_available_flights: {e}")
        await update.message.reply_text(f"Error retrieving flights: {e}")

# Admin functions
async def add_flight_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start the process of adding a new flight (admin only)."""
    ensure_excel_files_exist()
    
    # Check if user is admin
    if update.effective_user.id not in ADMIN_USERS:
        await update.message.reply_text("This command is only available to administrators.")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "Please enter the flight details in the following format:\n\n"
        "YYYY-MM-DD HH:MM FlightNumber Departure Destination Capacity\n\n"
        "Example: 2025-04-01 08:30 FL123 New_York London 120\n\n"
        "Type /cancel to cancel."
    )
    
    return ADMIN_ADDING_FLIGHT

async def add_flight_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Process new flight details from admin."""
    if update.message.text == '/cancel':
        await update.message.reply_text("Flight addition cancelled.")
        return ConversationHandler.END
    
    try:
        # Parse the input
        parts = update.message.text.split()
        if len(parts) < 6:
            await update.message.reply_text("Not enough details provided. Please try again with all required information.")
            return ADMIN_ADDING_FLIGHT
        
        date_str = parts[0]
        time_str = parts[1]
        flight_number = parts[2]
        departure = parts[3].replace('_', ' ')
        destination = parts[4].replace('_', ' ')
        capacity = int(parts[5])
        
        # Validate date
        try:
            flight_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            today = datetime.now().date()
            if flight_date < today:
                await update.message.reply_text("Cannot add flights for past dates.")
                return ADMIN_ADDING_FLIGHT
        except ValueError:
            await update.message.reply_text("Invalid date format. Please use YYYY-MM-DD.")
            return ADMIN_ADDING_FLIGHT
        
        # Read existing flights
        flights_df = pd.read_excel(FLIGHTS_FILE)
        
        # Check if flight already exists
        flights_df['Date'] = pd.to_datetime(flights_df['Date'])
        existing_flight = flights_df[
            (flights_df['Date'].dt.strftime("%Y-%m-%d") == date_str) & 
            (flights_df['Time'] == time_str) & 
            (flights_df['Flight Number'] == flight_number)
        ]
        
        if not existing_flight.empty:
            await update.message.reply_text("This flight already exists in the system.")
            return ADMIN_ADDING_FLIGHT
        
        # Add new flight
        new_flight = pd.DataFrame({
            'Date': [date_str],
            'Time': [time_str],
            'Flight Number': [flight_number],
            'Departure': [departure],
            'Destination': [destination],
            'Capacity': [capacity],
            'Booked': [0]
        })
        
        flights_df = pd.concat([flights_df, new_flight], ignore_index=True)
        flights_df.to_excel(FLIGHTS_FILE, index=False)
        
        await update.message.reply_text(
            f"✅ Flight added successfully:\n\n"
            f"Date: {date_str}\n"
            f"Time: {time_str}\n"
            f"Flight: {flight_number}\n"
            f"Route: {departure} to {destination}\n"
            f"Capacity: {capacity} seats"
        )
        
        return ConversationHandler.END
    
    except Exception as e:
        logger.error(f"Error in add_flight_finish: {e}")
        await update.message.reply_text(f"Error adding flight: {e}")
        return ADMIN_ADDING_FLIGHT

async def remove_flight(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove a flight (admin only)."""
    ensure_excel_files_exist()
    
    # Check if user is admin
    if update.effective_user.id not in ADMIN_USERS:
        await update.message.reply_text("This command is only available to administrators.")
        return
    
    if not context.args or len(context.args) < 3:
        await update.message.reply_text(
            "Please provide flight details: /remove_flight YYYY-MM-DD HH:MM FL123\n"
            "Example: /remove_flight 2025-04-01 08:30 FL123"
        )
        return
    
    try:
        date_str = context.args[0]
        time_str = context.args[1]
        flight_number = context.args[2]
        
        # Read flights data
        flights_df = pd.read_excel(FLIGHTS_FILE)
        flights_df['Date'] = pd.to_datetime(flights_df['Date'])
        
        # Find flight
        flight_to_remove = flights_df[
            (flights_df['Date'].dt.strftime("%Y-%m-%d") == date_str) & 
            (flights_df['Time'] == time_str) & 
            (flights_df['Flight Number'] == flight_number)
        ]
        
        if flight_to_remove.empty:
            await update.message.reply_text("Flight not found. Please check the details and try again.")
            return
        
        flight_index = flight_to_remove.index[0]
        
        # Check if there are bookings
        if flights_df.at[flight_index, 'Booked'] > 0:
            await update.message.reply_text(
                "⚠️ Warning: This flight has existing bookings. "
                "Please cancel these bookings first or use /force_remove_flight to remove it anyway."
            )
            return
        
        # Remove flight
        flights_df = flights_df.drop(flight_index)
        flights_df.to_excel(FLIGHTS_FILE, index=False)
        
        await update.message.reply_text(
            f"✅ Flight removed successfully:\n\n"
            f"Date: {date_str}\n"
            f"Time: {time_str}\n"
            f"Flight: {flight_number}"
        )
    
    except Exception as e:
        logger.error(f"Error in remove_flight: {e}")
        await update.message.reply_text(f"Error removing flight: {e}")

async def download_bookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Download bookings data (admin only)."""
    ensure_excel_files_exist()
    
    # Check if user is admin
    if update.effective_user.id not in ADMIN_USERS:
        await update.message.reply_text("This command is only available to administrators.")
        return
    
    try:
        await update.message.reply_document(
            document=open(BOOKINGS_FILE, 'rb'),
            filename=BOOKINGS_FILE
        )
    except Exception as e:
        logger.error(f"Error in download_bookings: {e}")
        await update.message.reply_text(f"Error sending bookings file: {e}")

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancel current conversation."""
    await update.message.reply_text("Operation cancelled.")
    return ConversationHandler.END

def main():
    """Start the bot."""
    # Create the Application
    application = Application.builder().token("7800384128:AAFz1oFxkJfCojXgQ1KDve5i3XCshjqPhak").build()
    
    # Add conversation handler for booking
    booking_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("book", book_flight_start)],
        states={
            SELECTING_DATE: [CallbackQueryHandler(select_date)],
            SELECTING_FLIGHT: [CallbackQueryHandler(select_flight)],
            CONFIRMING_BOOKING: [CallbackQueryHandler(confirm_booking)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Add conversation handler for cancelling bookings
    cancel_booking_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("cancel_booking", cancel_booking_start)],
        states={
            CONFIRMING_BOOKING: [CallbackQueryHandler(cancel_booking_confirm)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Add conversation handler for adding flights (admin only)
    add_flight_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("add_flight", add_flight_start)],
        states={
            ADMIN_ADDING_FLIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_flight_finish)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Add command handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("my_bookings", my_bookings))
    application.add_handler(CommandHandler("available", show_available_flights))
    application.add_handler(CommandHandler("remove_flight", remove_flight))
    application.add_handler(CommandHandler("download_bookings", download_bookings))
    application.add_handler(CommandHandler("recreate_excel", recreate_excel_files))
    
    application.add_handler(booking_conv_handler)
    application.add_handler(cancel_booking_conv_handler)
    application.add_handler(add_flight_conv_handler)
    
    # Run the bot until the user presses Ctrl-C
    application.run_polling()

if __name__ == "__main__":
    main()